"""数据中心查询 Excel 导出测试。"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.schemas.common import DataQueryParams
from app.services.data_export import DataExportService


@pytest.fixture
def export_svc(monkeypatch):
    from app.services.data_export import DataQueryService

    class _FakeQuery:
        def query(self, params):
            if params.category == "balance":
                return {
                    "category": "balance",
                    "rows": [
                        {
                            "agency": "IEA",
                            "snapshot_month": "2026-05",
                            "supply_demand": "供",
                            "period": "2026Q1",
                            "value": 102.0,
                        }
                    ],
                    "total": 1,
                    "snapshot_month": "2026-05",
                }
            return {"category": "price", "series": [], "monthly_stats": [], "total": 0}

        def charts_for(self, params):
            return [
                {
                    "title": "机构供需预测对比",
                    "xAxis": "周期",
                    "yAxis": "百万桶/天",
                    "dual_y": False,
                    "series": [
                        {
                            "name": "IEA 供",
                            "data": [["2026Q1", 102.0]],
                            "yAxisIndex": 0,
                        }
                    ],
                }
            ]

    svc = DataExportService(db=None)  # type: ignore[arg-type]
    svc.query_svc = _FakeQuery()
    return svc


def test_export_balance_xlsx_has_sheets_and_chart(export_svc, tmp_path, monkeypatch):
    monkeypatch.setattr(
        "app.services.data_export.settings.exports_dir",
        tmp_path,
    )
    params = DataQueryParams(category="balance", year=2026, month=5)
    out = export_svc.export_query_xlsx(params)
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert "查询条件" in wb.sheetnames
    assert "供需" in wb.sheetnames

    ws = wb["供需"]
    values = [ws.cell(r, 1).value for r in range(1, 20)]
    assert any(v == "供需预测数据" for v in values)
    assert ws._images, "应嵌入图表图片"
