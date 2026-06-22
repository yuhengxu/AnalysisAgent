"""预测分析表服务：生成（调用 skill）、增删改查、导出 Excel。"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.core import llm
from app.core.config import settings
from app.core.llm import llm_context
from app.core.llm_models import get_default_model_for_provider
from app.core.timezone import beijing_timestamp, format_beijing_iso, now_beijing_naive
from app.models.prediction import Prediction
from app.models.user import User
from app.skills.prediction_skill import PredictionSkill
from app.skills.prediction_unrestricted_skill import PredictionUnrestrictedSkill
from app.templates.prediction_table import (
    CONFIDENCE_LEVELS,
    IMPACT_OPTIONS,
    PREDICTION_FACTORS,
    PREDICTION_NOTES,
)


class PredictionService:
    def __init__(self, db: Session):
        self.db = db
        self.skill = PredictionSkill(db)
        self.unrestricted_skill = PredictionUnrestrictedSkill(db)

    _FACTOR_EXPERT_OPINION_KEYS = (
        "revise_opinions",
        "expert_opinion",
        "expert_opinions",
        "importance_opinion",
        "judgment_opinion",
        "impact_opinion",
    )
    _CONTENT_EXPERT_OPINION_KEYS = ("revise_opinions", "expert_opinions")

    @classmethod
    def _strip_expert_opinions(cls, content: dict[str, Any]) -> dict[str, Any]:
        """移除专家修改意见字段，不入库。"""
        stripped = json.loads(json.dumps(content, ensure_ascii=False))
        for key in cls._CONTENT_EXPERT_OPINION_KEYS:
            stripped.pop(key, None)
        for factor in stripped.get("factors", []):
            for key in cls._FACTOR_EXPERT_OPINION_KEYS:
                factor.pop(key, None)
        return stripped

    @classmethod
    def _strip_content_for_export(cls, content: dict[str, Any]) -> dict[str, Any]:
        """导出 Excel 前移除仅用于前端展示的字段（含专家修改意见）。"""
        stripped = cls._strip_expert_opinions(content)
        for factor in stripped.get("factors", []):
            factor.pop("source_refs", None)
            factor.pop("confidence_level", None)
            factor.pop("source_url", None)
            factor.pop("source_title", None)
        return stripped

    # ------------------------------------------------------------------ #
    @staticmethod
    def _scope_query(query, user: User | None):
        if user is None:
            return query.filter(False)
        return query.filter(Prediction.user_id == user.id)

    def get(self, pred_id: int, user: User | None = None) -> Prediction | None:
        p = self.db.get(Prediction, pred_id)
        if not p:
            return None
        if user is None:
            return None
        if p.user_id != user.id:
            return None
        return p

    def generate(
        self,
        symbol: str = "Brent",
        year: int = 2026,
        month: int = 6,
        provider: str | None = None,
        model: str | None = None,
        mode: str = "deep_research",
        extra_instruction: str = "",
        on_progress: Callable[[int, int, str], None] | None = None,
        trusted_sources_only: bool = False,
        unrestricted_mode: bool = False,
        user: User | None = None,
    ) -> dict[str, Any]:
        if unrestricted_mode:
            result = self.unrestricted_skill.generate(
                symbol=symbol,
                year=year,
                month=month,
                provider=provider,
                model=model,
                mode=mode,
                extra_instruction=extra_instruction,
                on_progress=on_progress,
            )
        else:
            result = self.skill.generate(
                symbol=symbol,
                year=year,
                month=month,
                provider=provider,
                model=model,
                mode=mode,
                extra_instruction=extra_instruction,
                on_progress=on_progress,
                trusted_sources_only=trusted_sources_only,
            )
        evidence = dict(result["evidence"])
        if result.get("web_references"):
            evidence["web_references"] = result["web_references"]
        pred = Prediction(
            title=f"{year}年{month}月{symbol}油价预测分析表",
            symbol=symbol,
            year=year,
            month=month,
            status="draft",
            user_id=user.id if user else None,
            content_json=json.dumps(result["content"], ensure_ascii=False),
            evidence_json=json.dumps(evidence, ensure_ascii=False, default=str),
            sources_json=json.dumps(result["sources_used"], ensure_ascii=False),
            model_name=result.get("model", ""),
            llm_used=1 if result.get("llm_used") else 0,
        )
        self.db.add(pred)
        self.db.commit()
        self.db.refresh(pred)
        detail = self.get_detail(pred.id, user)
        detail["total_steps"] = result.get("total_steps", 7)
        return detail

    # ------------------------------------------------------------------ #
    def list_predictions(self, user: User | None = None) -> list[dict[str, Any]]:
        query = self.db.query(Prediction).order_by(Prediction.created_at.desc())
        rows = self._scope_query(query, user).all()
        return [
            {
                "id": r.id,
                "title": r.title,
                "symbol": r.symbol,
                "year": r.year,
                "month": r.month,
                "status": r.status,
                "llm_used": bool(r.llm_used),
                "model_name": r.model_name,
                "updated_at": format_beijing_iso(r.updated_at),
            }
            for r in rows
        ]

    def get_detail(self, pred_id: int, user: User | None = None) -> dict[str, Any]:
        p = self.get(pred_id, user)
        if not p:
            raise ValueError("Prediction not found")
        return {
            "id": p.id,
            "title": p.title,
            "symbol": p.symbol,
            "year": p.year,
            "month": p.month,
            "status": p.status,
            "content": json.loads(p.content_json),
            "evidence": json.loads(p.evidence_json or "{}"),
            "sources": json.loads(p.sources_json or "[]"),
            "llm_used": bool(p.llm_used),
            "model_name": p.model_name,
            "template": {
                "categories": PREDICTION_FACTORS,
                "impact_options": IMPACT_OPTIONS,
                "confidence_levels": CONFIDENCE_LEVELS,
                "notes": PREDICTION_NOTES,
            },
        }

    def update_content(
        self,
        pred_id: int,
        content: dict[str, Any],
        title: str | None = None,
        user: User | None = None,
    ) -> dict[str, Any]:
        p = self.get(pred_id, user)
        if not p:
            raise ValueError("Prediction not found")
        p.content_json = json.dumps(self._strip_expert_opinions(content), ensure_ascii=False)
        if title:
            p.title = title
        p.status = "reviewed"
        p.updated_at = now_beijing_naive()
        self.db.commit()
        return {"id": p.id, "status": "updated"}

    def revise_factor_field(
        self,
        pred_id: int,
        factor_idx: int,
        field: str,
        instruction: str,
        provider: str | None = None,
        model: str | None = None,
        mode: str = "deep_research",
        user: User | None = None,
    ) -> dict[str, Any]:
        p = self.get(pred_id, user)
        if not p:
            raise ValueError("Prediction not found")
        content = json.loads(p.content_json)
        factors = content.get("factors", [])
        if factor_idx < 0 or factor_idx >= len(factors):
            raise ValueError("Factor index out of range")
        factor = factors[factor_idx]
        if field != "judgment":
            raise ValueError("Unsupported field")

        old_value = factor.get(field, "")
        revised = self._llm_revise_factor_field(
            factor_name=f"{factor.get('id', '')} {factor.get('name', '')}".strip(),
            field=field,
            old_value=old_value,
            instruction=instruction,
            provider=provider,
            model=model,
            mode=mode,
        )
        factor[field] = revised
        p.content_json = json.dumps(content, ensure_ascii=False)
        p.status = "reviewed"
        p.updated_at = now_beijing_naive()
        self.db.commit()
        return {"factor_idx": factor_idx, "field": field, "value": revised}

    def _llm_revise_factor_field(
        self,
        *,
        factor_name: str,
        field: str,
        old_value: Any,
        instruction: str,
        provider: str | None,
        model: str | None,
        mode: str = "deep_research",
    ) -> Any:
        prov = provider or settings.default_llm_provider
        resolved_model = model or get_default_model_for_provider(prov)
        field_labels = {"judgment": "形势判断及支撑指标"}
        field_rules = {
            "judgment": "只输出改写后的形势判断正文，保留数据与来源标注，不要解释过程。",
        }
        if llm.is_enabled(prov):
            try:
                with llm_context(
                    "prediction_revise",
                    field=field,
                    factor=factor_name[:80],
                    instruction=instruction[:200],
                    mode=mode,
                ):
                    raw = llm.chat(
                        [
                            {
                                "role": "system",
                                "content": (
                                    "你是国际油价预测分析表的资深专家编辑。"
                                    "请根据专家修改意见，修订指定因素的某一列内容。"
                                    f"{field_rules[field]}"
                                ),
                            },
                            {
                                "role": "user",
                                "content": (
                                    f"影响因素：{factor_name}\n"
                                    f"列名：{field_labels[field]}\n"
                                    f"当前内容：{old_value}\n\n"
                                    f"专家修改意见：{instruction}"
                                ),
                            },
                        ],
                        provider=prov,
                        model=resolved_model,
                        temperature=0.3,
                        mode=mode,
                    ).strip()
                return self._parse_revised_field(field, raw, old_value)
            except llm.LLMUnavailable:
                pass
        return old_value

    @staticmethod
    def _parse_revised_field(field: str, raw: str, fallback: Any) -> Any:
        del field
        return raw.strip() or fallback

    def delete(self, pred_id: int, user: User | None = None) -> None:
        p = self.get(pred_id, user)
        if not p:
            raise ValueError("Prediction not found")
        if p.xlsx_path:
            fp = Path(p.xlsx_path)
            if fp.exists():
                fp.unlink()
        self.db.delete(p)
        self.db.commit()

    # ------------------------------------------------------------------ #
    def export_xlsx(self, pred_id: int, user: User | None = None) -> Path:
        p = self.get(pred_id, user)
        if not p:
            raise ValueError("Prediction not found")
        content = self._strip_content_for_export(json.loads(p.content_json))
        path = self._build_xlsx(p, content)
        p.xlsx_path = str(path)
        self.db.commit()
        return path

    def _build_xlsx(self, p: Prediction, content: dict[str, Any]) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{p.month}月国际油价预测"

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        cat_fill = PatternFill("solid", fgColor="DDEBF7")
        wrap = Alignment(wrap_text=True, vertical="center")
        center = Alignment(horizontal="center", vertical="center")

        # 列宽
        widths = {"B": 26, "C": 5, "D": 5, "E": 5, "F": 5, "G": 5, "H": 70, "I": 7, "J": 7, "K": 7}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # 标题（对齐 yuebao/prediction 样例：B3）
        ws.merge_cells("B3:K3")
        c = ws["B3"]
        c.value = "近期国际油价预测分析表"
        c.font = Font(size=14, bold=True)
        c.alignment = center

        # 表头
        ws.merge_cells("C5:G5")
        ws["B5"] = "影响因素"
        ws["C5"] = "重要性程度\n（1为最低，5为最高，\n请在数字下划线）"
        ws["H5"] = "形势判断及支撑指标"
        ws.merge_cells("I5:K5")
        ws["I5"] = "对国际油价影响"
        for col in ("B", "C", "H", "I"):
            cell = ws[f"{col}5"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center

        factors_by_id = {f["id"]: f for f in content.get("factors", [])}
        row = 6
        for cat in PREDICTION_FACTORS:
            ws.merge_cells(f"B{row}:K{row}")
            cell = ws[f"B{row}"]
            cell.value = cat["title"]
            cell.font = Font(bold=True)
            cell.fill = cat_fill
            cell.alignment = Alignment(vertical="center")
            row += 1
            for fdef in cat["factors"]:
                data = factors_by_id.get(fdef["id"], {})
                ws[f"B{row}"] = f"  {fdef['id']} {fdef['name']}"
                ws[f"B{row}"].alignment = wrap
                importance = int(data.get("importance", 1) or 1)
                for i, col in enumerate(("C", "D", "E", "F", "G"), start=1):
                    cell = ws[f"{col}{row}"]
                    cell.value = i
                    cell.alignment = center
                    if i == importance:
                        cell.font = Font(underline="single", bold=True)
                ws[f"H{row}"] = data.get("judgment", "")
                ws[f"H{row}"].alignment = wrap
                impact = data.get("impact", "持平")
                for col, opt in (("I", "促涨"), ("J", "持平"), ("K", "促跌")):
                    cell = ws[f"{col}{row}"]
                    cell.value = opt
                    cell.alignment = center
                    if opt == impact:
                        cell.font = Font(underline="single", bold=True)
                row += 1

        # 价格预测
        pf = content.get("price_forecast", {})
        for key in ("current_month", "next_month"):
            blk = pf.get(key, {}) or {}
            ws.merge_cells(f"B{row}:G{row}")
            ws[f"B{row}"] = blk.get("label", "")
            ws[f"B{row}"].font = Font(bold=True)
            lo, hi = blk.get("range_low"), blk.get("range_high")
            ws.merge_cells(f"H{row}:H{row}")
            ws[f"H{row}"] = f"区间（幅度不超过5美元）：{lo}-{hi}" if lo is not None else "区间："
            ws.merge_cells(f"I{row}:K{row}")
            ws[f"I{row}"] = f"均价：{blk.get('avg')}" if blk.get("avg") is not None else "均价："
            ws[f"I{row}"].alignment = center
            row += 1

        # 注
        for note in PREDICTION_NOTES:
            ws.merge_cells(f"B{row}:K{row}")
            ws[f"B{row}"] = note
            ws[f"B{row}"].font = Font(size=9, color="808080")
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        # 边框
        for r in range(5, row):
            for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K"):
                ws[f"{col}{r}"].border = border

        out = settings.exports_dir / (
            f"prediction_{p.id}_{beijing_timestamp()}.xlsx"
        )
        wb.save(out)
        return out
