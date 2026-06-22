"""数据中心查询结果导出 Excel（含表格与图表）。"""
from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.timezone import beijing_timestamp
from app.schemas.common import DataQueryParams
from app.services.chart_matplotlib import render_report_chart_matplotlib
from app.services.data_query import DataQueryService

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)


class DataExportService:
    def __init__(self, db: Session):
        self.db = db
        self.query_svc = DataQueryService(db)

    def export_query_xlsx(self, params: DataQueryParams) -> Path:
        export_params = params.model_copy(update={"page": 1, "page_size": 50_000})
        data = self.query_svc.query(export_params)
        charts = self.query_svc.charts_for(params)

        settings.exports_dir.mkdir(parents=True, exist_ok=True)
        stamp = beijing_timestamp()
        out = settings.exports_dir / f"data_query_{params.category}_{stamp}.xlsx"

        with tempfile.TemporaryDirectory(prefix="data_export_") as tmp:
            work_dir = Path(tmp)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            self._write_params_sheet(wb, params)
            if params.category == "mixed":
                self._write_mixed_sheets(wb, data, charts, work_dir)
            elif params.category == "price":
                self._write_price_sheet(wb, data, charts, work_dir)
            elif params.category == "balance":
                self._write_balance_sheet(wb, data, charts, work_dir)
            else:
                self._write_factor_sheet(wb, data, charts, work_dir)
            wb.save(out)
        return out

    def _write_params_sheet(self, wb: openpyxl.Workbook, params: DataQueryParams) -> None:
        ws = wb.create_sheet("查询条件", 0)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 64
        rows: list[tuple[str, Any]] = [
            ("数据类别", params.category),
            ("起始日期", params.start_date),
            ("结束日期", params.end_date),
            ("年月", f"{params.year}-{params.month:02d}" if params.year and params.month else ""),
            ("品种", ", ".join(params.symbols)),
            ("机构", ", ".join(params.agencies)),
            ("供需类型", ", ".join(params.supply_demand)),
            ("周期", ", ".join(params.periods)),
            ("因素大类", ", ".join(params.factor_categories)),
            ("因素名", ", ".join(params.factor_names)),
        ]
        ws["A1"] = "数据中心查询导出"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:B1")
        start = 3
        for idx, (label, value) in enumerate(rows):
            r = start + idx
            ws.cell(r, 1, label).font = Font(bold=True)
            ws.cell(r, 2, "" if value is None else str(value))

    def _write_mixed_sheets(
        self,
        wb: openpyxl.Workbook,
        data: dict[str, Any],
        charts: list[dict[str, Any]],
        work_dir: Path,
    ) -> None:
        price_charts = [c for c in charts if _is_price_chart(c)]
        balance_charts = [c for c in charts if _is_balance_chart(c)]
        if data.get("price"):
            self._write_price_sheet(
                wb,
                {"category": "price", **data["price"]},
                price_charts or [c for c in charts if c not in balance_charts],
                work_dir,
            )
        if data.get("balance"):
            self._write_balance_sheet(
                wb,
                {"category": "balance", **data["balance"]},
                balance_charts or [c for c in charts if c not in price_charts],
                work_dir,
            )
        if data.get("factor"):
            self._write_factor_sheet(wb, {"category": "factor", **data["factor"]}, [], work_dir)

    def _write_price_sheet(
        self,
        wb: openpyxl.Workbook,
        data: dict[str, Any],
        charts: list[dict[str, Any]],
        work_dir: Path,
    ) -> None:
        ws = wb.create_sheet("价格")
        row = 1
        stats = data.get("monthly_stats") or []
        if stats:
            ws.cell(row, 1, "月度统计").font = TITLE_FONT
            row = _write_dict_table(ws, stats, row + 1) + 2

        series = data.get("series") or []
        ws.cell(row, 1, "日度序列").font = TITLE_FONT
        note = ""
        export_series = series
        if len(series) > 5000:
            export_series = series[:5000]
            note = f"仅导出前 5000 条，共 {len(series)} 条"
        row = _write_dict_table(ws, export_series, row + 1) + 1
        if note:
            ws.cell(row, 1, note)
            row += 2

        for idx, chart in enumerate(charts):
            row = _append_chart_section(ws, chart, work_dir, row, idx)

    def _write_balance_sheet(
        self,
        wb: openpyxl.Workbook,
        data: dict[str, Any],
        charts: list[dict[str, Any]],
        work_dir: Path,
    ) -> None:
        ws = wb.create_sheet("供需")
        row = 1
        snapshot = data.get("snapshot_month")
        if snapshot:
            ws.cell(row, 1, f"快照月份：{snapshot}")
            row += 2
        ws.cell(row, 1, "供需预测数据").font = TITLE_FONT
        row = _write_dict_table(ws, data.get("rows") or [], row + 1) + 2
        for idx, chart in enumerate(charts):
            row = _append_chart_section(ws, chart, work_dir, row, idx)

    def _write_factor_sheet(
        self,
        wb: openpyxl.Workbook,
        data: dict[str, Any],
        charts: list[dict[str, Any]],
        work_dir: Path,
    ) -> None:
        ws = wb.create_sheet("预测因素")
        row = 1
        report_month = data.get("report_month")
        if report_month:
            ws.cell(row, 1, f"报告月份：{report_month}")
            row += 2
        ws.cell(row, 1, "因素评估数据").font = TITLE_FONT
        row = _write_dict_table(ws, data.get("rows") or [], row + 1) + 2
        for idx, chart in enumerate(charts):
            row = _append_chart_section(ws, chart, work_dir, row, idx)


def _is_price_chart(chart: dict[str, Any]) -> bool:
    title = str(chart.get("title", ""))
    return "供需" not in title and "机构" not in title


def _is_balance_chart(chart: dict[str, Any]) -> bool:
    title = str(chart.get("title", ""))
    return "供需" in title or "机构" in title


def _write_dict_table(ws, rows: list[dict[str, Any]], start_row: int) -> int:
    if not rows:
        ws.cell(start_row, 1, "（无数据）")
        return start_row + 1
    headers = list(rows[0].keys())
    for ci, header in enumerate(headers, 1):
        cell = ws.cell(start_row, ci, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        width = max(12, min(40, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(ci)].width = width
    for ri, row in enumerate(rows, start_row + 1):
        for ci, header in enumerate(headers, 1):
            val = row.get(header)
            ws.cell(ri, ci, val if val is not None else "")
    return start_row + len(rows)


def _chart_series_table(chart: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for serie in chart.get("series") or []:
        name = serie.get("name", "")
        for x, y in serie.get("data") or []:
            rows.append({"系列": name, "X": x, "Y": y})
    return rows


def _append_chart_section(ws, chart: dict[str, Any], work_dir: Path, start_row: int, chart_idx: int = 0) -> int:
    title = str(chart.get("title") or "图表")
    ws.cell(start_row, 1, title).font = TITLE_FONT
    row = start_row + 1

    chart_rows = _chart_series_table(chart)
    if chart_rows:
        row = _write_dict_table(ws, chart_rows, row) + 1

    png = work_dir / f"chart_{chart_idx}.png"
    if render_report_chart_matplotlib(chart, png) and png.exists():
        img = XLImage(str(png))
        img.width = min(640, img.width or 640)
        img.height = min(360, img.height or 335)
        anchor = f"A{row}"
        ws.add_image(img, anchor)
        row += 22
    else:
        ws.cell(row, 1, "（图表渲染失败）")
        row += 2
    return row + 1
