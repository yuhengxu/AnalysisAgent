import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.timezone import beijing_timestamp
from app.models.balance_forecast import BalanceForecast
from app.models.dataset import Dataset
from app.models.factor_assessment import FactorAssessment
from app.models.price_series import PriceSeries
from app.templates.sample_contracts import (
    BALANCE_AGENCIES,
    BALANCE_HEADER_NAMES,
    BALANCE_SUPPLY_DEMAND_ALIASES,
    BALANCE_SUPPLY_DEMAND_VALUES,
    PRICE_HEADER_MARKERS,
    is_sample_aggregate_label,
)
from app.templates.monthly_report import MONTHLY_PRICE_SOURCE, PRICE_SYMBOL_MAP
_MONTH_AVG_LABEL_RE = re.compile(r"^(\d{2})(\d{2})-(\d{2})(\d{2})平均$")
_SHEET_PERIOD_RE = re.compile(r"^(\d{4})年(\d{1,2})月$")
_BALANCE_SHEET_RE = re.compile(r"^(\d{4})(\d{2})$")
_PERIOD_COL_RE = re.compile(r"20\d{2}(?:Q[1-4]|年)")
_FACTOR_MONTH_PATTERNS = (
    re.compile(r"(\d{4})年(\d{1,2})月"),
    re.compile(r"(\d{4})[-_/](\d{1,2})"),
    re.compile(r"(\d{4})(\d{2})(?!\d)"),
)


class ImportCounter:
    def __init__(self) -> None:
        self.inserted = 0
        self.updated = 0
        self.skipped = 0
        self.errors: list[str] = []
        self.imported_sheets: list[str] = []
        self.skipped_sheets: list[dict[str, str]] = []

    @property
    def affected(self) -> int:
        return self.inserted + self.updated

    def skip_sheet(self, sheet: str, reason: str) -> None:
        self.skipped += 1
        self.skipped_sheets.append({"sheet": sheet, "reason": reason})

    def to_dict(self) -> dict[str, Any]:
        return {
            "inserted": self.inserted,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": self.errors,
            "imported_sheets": self.imported_sheets,
            "skipped_sheets": self.skipped_sheets,
        }


class DataImportService:
    def __init__(self, db: Session):
        self.db = db

    def save_upload(self, filename: str, content: bytes) -> Path:
        dest = settings.raw_dir / f"{beijing_timestamp()}_{filename}"
        dest.write_bytes(content)
        return dest

    def detect_category(self, filename: str) -> str:
        name = filename.lower()
        if "价格" in filename or "price" in name:
            return "price"
        if "供需" in filename or "balance" in name:
            return "balance"
        if "预测" in filename or "factor" in name:
            return "factor"
        return "generic"

    def import_file(self, file_path: Path, category: str | None = None) -> dict[str, Any]:
        category = category or self.detect_category(file_path.name)
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._import_csv(file_path, category)
        if suffix in {".xlsx", ".xls"}:
            return self._import_xlsx(file_path, category)
        raise ValueError(f"Unsupported file type: {suffix}")

    def _create_dataset(self, name: str, category: str, file_path: Path, schema: dict, row_count: int) -> Dataset:
        ds = Dataset(
            name=name,
            source_type=file_path.suffix.lstrip("."),
            category=category,
            file_path=str(file_path),
            schema_json=json.dumps(schema, ensure_ascii=False),
            row_count=row_count,
        )
        self.db.add(ds)
        self.db.commit()
        self.db.refresh(ds)
        return ds

    def _import_csv(self, file_path: Path, category: str) -> dict[str, Any]:
        df = pd.read_csv(file_path)
        ds = self._create_dataset(file_path.stem, category, file_path, {"columns": list(df.columns)}, len(df))
        return {"dataset_id": ds.id, "category": category, "rows": len(df), "message": "CSV imported as dataset metadata"}

    def _import_xlsx(self, file_path: Path, category: str) -> dict[str, Any]:
        if category == "price":
            return self._import_price_xlsx(file_path)
        if category == "balance":
            return self._import_balance_xlsx(file_path)
        if category == "factor":
            return self._import_factor_xlsx(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ds = self._create_dataset(file_path.stem, category, file_path, {"sheets": wb.sheetnames}, 0)
        wb.close()
        return {"dataset_id": ds.id, "category": category, "sheets": wb.sheetnames}

    def _import_price_xlsx(self, file_path: Path) -> dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        counter = ImportCounter()
        sheet_meta: list[dict[str, Any]] = []
        ds = self._create_dataset(file_path.stem, "price", file_path, {"sheets": wb.sheetnames}, 0)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not _SHEET_PERIOD_RE.match(str(sheet_name).strip()):
                counter.skip_sheet(sheet_name, "sheet_name_not_YYYY年M月")
                continue
            header_info = self._detect_price_header(rows)
            if not header_info:
                counter.skip_sheet(sheet_name, "price_header_not_found")
                continue
            header_idx, date_col, symbol_cols = header_info
            sheet_rows = 0
            sheet_period = self._parse_sheet_period(sheet_name)
            for row in rows[header_idx + 1 :]:
                if not row or date_col >= len(row) or row[date_col] is None:
                    counter.skipped += 1
                    continue
                first_cell = str(row[date_col]).strip()
                if sheet_period and self._is_monthly_avg_label_for_period(first_cell, sheet_period[1]):
                    month_rows = self._import_monthly_avg_row(
                        dataset_id=ds.id,
                        row=row,
                        date_col=date_col,
                        symbol_cols=symbol_cols,
                        year=sheet_period[0],
                        month=sheet_period[1],
                    )
                    counter.inserted += month_rows["inserted"]
                    counter.updated += month_rows["updated"]
                    counter.skipped += month_rows["skipped"]
                    sheet_rows += month_rows["affected"]
                    continue
                if is_sample_aggregate_label(first_cell):
                    counter.skipped += 1
                    continue
                date_val = self._parse_date(row[date_col])
                if not date_val:
                    counter.skipped += 1
                    continue
                seen_symbols_in_row: set[str] = set()
                for idx, symbol in symbol_cols.items():
                    if symbol in seen_symbols_in_row:
                        counter.skipped += 1
                        continue
                    seen_symbols_in_row.add(symbol)
                    if idx >= len(row):
                        continue
                    price = row[idx]
                    if price is None or str(price).strip() == "":
                        continue
                    try:
                        price_f = float(price)
                    except (TypeError, ValueError):
                        counter.skipped += 1
                        continue
                    changed = self._upsert_price(
                        dataset_id=ds.id,
                        symbol=symbol,
                        trade_date=date_val,
                        price=price_f,
                        source="CNEEI",
                    )
                    if changed == "inserted":
                        counter.inserted += 1
                    elif changed == "updated":
                        counter.updated += 1
                    else:
                        counter.skipped += 1
                    sheet_rows += 1
            sheet_meta.append(
                {
                    "sheet": sheet_name,
                    "header_row": header_idx + 1,
                    "date_col": date_col + 1,
                    "symbols": list(symbol_cols.values()),
                    "rows": sheet_rows,
                }
            )
            counter.imported_sheets.append(sheet_name)

        if not counter.imported_sheets:
            self.db.rollback()
            wb.close()
            raise ValueError(
                json.dumps({"message": "无有效价格 sheet", "skipped_sheets": counter.skipped_sheets}, ensure_ascii=False)
            )

        ds.row_count = counter.affected
        ds.schema_json = json.dumps({"sheets": sheet_meta}, ensure_ascii=False, default=str)
        self.db.commit()
        wb.close()
        return {"dataset_id": ds.id, "category": "price", "rows": counter.affected, **counter.to_dict()}

    def _import_balance_xlsx(self, file_path: Path) -> dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ds = self._create_dataset(file_path.stem, "balance", file_path, {"sheets": wb.sheetnames}, 0)
        counter = ImportCounter()
        sheet_meta: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            snapshot_month = self._parse_balance_snapshot_month(sheet_name)
            if not snapshot_month:
                counter.skip_sheet(sheet_name, "sheet_name_not_YYYYMM")
                continue
            parsed = self._parse_balance_sheet(rows, sheet_name, snapshot_month)
            if not parsed:
                counter.skip_sheet(sheet_name, "layout_unrecognized")
                continue
            sheet_rows = 0
            for item in parsed["records"]:
                changed = self._upsert_balance(dataset_id=ds.id, **item)
                if changed == "inserted":
                    counter.inserted += 1
                elif changed == "updated":
                    counter.updated += 1
                else:
                    counter.skipped += 1
                sheet_rows += 1
            if sheet_rows == 0:
                counter.skip_sheet(sheet_name, "no_valid_rows")
                continue
            sheet_meta.append({"sheet": sheet_name, "snapshot_month": snapshot_month, "layout": parsed["layout"], "rows": sheet_rows})
            counter.imported_sheets.append(sheet_name)

        if not counter.imported_sheets:
            self.db.rollback()
            wb.close()
            raise ValueError(
                json.dumps({"message": "无有效供需 sheet", "skipped_sheets": counter.skipped_sheets}, ensure_ascii=False)
            )

        ds.row_count = counter.affected
        ds.schema_json = json.dumps({"sheets": sheet_meta}, ensure_ascii=False, default=str)
        self.db.commit()
        wb.close()
        return {"dataset_id": ds.id, "category": "balance", "rows": counter.affected, **counter.to_dict()}

    @staticmethod
    def _parse_balance_snapshot_month(sheet_name: str) -> str | None:
        match = _BALANCE_SHEET_RE.fullmatch(str(sheet_name).strip())
        if not match:
            return None
        return f"{match.group(1)}-{match.group(2)}"

    @staticmethod
    def _normalize_supply_demand(value: str) -> str | None:
        text = str(value or "").strip()
        return BALANCE_SUPPLY_DEMAND_ALIASES.get(text)

    @staticmethod
    def _normalize_period_header(header: str) -> tuple[str, bool]:
        text = str(header).strip()
        if "供需差" in text:
            match = re.search(r"(20\d{2}Q[1-4]|20\d{2}年(?:\d{1,2}月)?)", text)
            return (match.group(1) if match else text.replace("供需差", "").strip()), True
        return text, False

    def _parse_balance_sheet(self, rows: list[tuple[Any, ...]], sheet_name: str, snapshot_month: str) -> dict[str, Any] | None:
        layout_c = self._detect_balance_layout_c(rows)
        if layout_c:
            return {"layout": "C", "records": self._parse_balance_layout_c(rows, sheet_name, snapshot_month, layout_c)}
        header_info = self._detect_balance_header(rows)
        if header_info:
            layout = "B" if header_info[1] > 1 else "A"
            return {
                "layout": layout,
                "records": self._parse_balance_layout_ab(rows, sheet_name, snapshot_month, header_info),
            }
        return None

    def _parse_balance_layout_ab(
        self,
        rows: list[tuple[Any, ...]],
        sheet_name: str,
        snapshot_month: str,
        header_info: tuple[int, int, int, int, list[tuple[int, str]]],
    ) -> list[dict[str, Any]]:
        header_idx, agency_col, update_col, sd_col, period_cols = header_info
        current_agency = ""
        current_update = ""
        records: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if not row or sd_col >= len(row):
                continue
            agency_cell = row[agency_col] if agency_col < len(row) else None
            update_cell = row[update_col] if update_col < len(row) else None
            sd_cell = row[sd_col] if sd_col < len(row) else None
            if agency_cell and str(agency_cell).strip():
                current_agency = self._normalize_agency(str(agency_cell).strip())
            if update_cell and str(update_cell).strip():
                current_update = self._format_update_date(update_cell, sheet_name)
            sd = self._normalize_supply_demand(str(sd_cell).strip() if sd_cell else "")
            if not current_agency or not sd:
                continue
            for idx, period_header in period_cols:
                if idx >= len(row) or row[idx] is None:
                    continue
                period, is_gap_col = self._normalize_period_header(period_header)
                if not _PERIOD_COL_RE.search(period):
                    continue
                try:
                    value = float(row[idx])
                except (TypeError, ValueError):
                    continue
                supply_demand = "供需差" if is_gap_col else sd
                records.append(
                    {
                        "agency": current_agency,
                        "snapshot_month": snapshot_month,
                        "update_date": current_update,
                        "supply_demand": supply_demand,
                        "period": period,
                        "value": value,
                        "balance_gap": value if supply_demand == "供需差" else None,
                    }
                )
        return records

    def _detect_balance_layout_c(
        self, rows: list[tuple[Any, ...]]
    ) -> tuple[int, int, int, list[tuple[int, str]]] | None:
        period_header: tuple[int, list[tuple[int, str]]] | None = None
        for r_idx, row in enumerate(rows[:20]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            period_cols = [(idx, col) for idx, col in enumerate(cells) if re.fullmatch(r"20\d{2}Q[1-4]", col)]
            if len(period_cols) >= 2 and BALANCE_HEADER_NAMES["supply_demand"] not in cells:
                period_header = (r_idx, period_cols)
                break
        if not period_header:
            return None
        header_idx, period_cols = period_header
        sd_col: int | None = None
        agency_col: int | None = None
        for row in rows[header_idx + 1 : header_idx + 12]:
            if not row:
                continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            if sd_col is None:
                found_sd = next((idx for idx, col in enumerate(cells) if col in BALANCE_SUPPLY_DEMAND_ALIASES), None)
                if found_sd is not None:
                    sd_col = found_sd
            for idx, cell in enumerate(cells):
                if self._normalize_agency(cell) in BALANCE_AGENCIES:
                    agency_col = idx
                    break
            if sd_col is not None and agency_col is not None:
                break
        if sd_col is None:
            return None
        if agency_col is None:
            agency_col = max(0, sd_col - 1)
        return header_idx, agency_col, sd_col, period_cols

    def _parse_balance_layout_c(
        self,
        rows: list[tuple[Any, ...]],
        sheet_name: str,
        snapshot_month: str,
        layout: tuple[int, int, int, list[tuple[int, str]]],
    ) -> list[dict[str, Any]]:
        header_idx, agency_col, sd_col, period_cols = layout
        current_agency = ""
        records: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if not row or sd_col >= len(row):
                continue
            agency_cell = row[agency_col] if agency_col < len(row) else None
            sd_cell = row[sd_col] if sd_col < len(row) else None
            if agency_cell and str(agency_cell).strip():
                normalized = self._normalize_agency(str(agency_cell).strip())
                if normalized in BALANCE_AGENCIES:
                    current_agency = normalized
            sd = self._normalize_supply_demand(str(sd_cell).strip() if sd_cell else "")
            if not current_agency or not sd:
                continue
            for idx, period in period_cols:
                if idx >= len(row) or row[idx] is None:
                    continue
                try:
                    value = float(row[idx])
                except (TypeError, ValueError):
                    continue
                records.append(
                    {
                        "agency": current_agency,
                        "snapshot_month": snapshot_month,
                        "update_date": "",
                        "supply_demand": sd,
                        "period": period,
                        "value": value,
                        "balance_gap": value if sd == "供需差" else None,
                    }
                )
        return records

    @staticmethod
    def _parse_sheet_period(sheet_name: str) -> tuple[int, int] | None:
        match = _SHEET_PERIOD_RE.match(str(sheet_name).strip())
        if not match:
            return None
        return int(match.group(1)), int(match.group(2))

    @staticmethod
    def _is_monthly_avg_label_for_period(label: str, sheet_month: int) -> bool:
        match = _MONTH_AVG_LABEL_RE.match(str(label).strip())
        if not match:
            return False
        return int(match.group(1)) == sheet_month

    @staticmethod
    def _month_last_day(year: int, month: int) -> date:
        if month == 12:
            return date(year, 12, 31)
        return date(year, month + 1, 1) - timedelta(days=1)

    def _import_monthly_avg_row(
        self,
        *,
        dataset_id: int,
        row: tuple[Any, ...],
        date_col: int,
        symbol_cols: dict[int, str],
        year: int,
        month: int,
    ) -> dict[str, int]:
        result = {"inserted": 0, "updated": 0, "skipped": 0, "affected": 0}
        trade_date = self._month_last_day(year, month)
        seen_symbols: set[str] = set()
        for idx, symbol in symbol_cols.items():
            if symbol in seen_symbols:
                result["skipped"] += 1
                continue
            seen_symbols.add(symbol)
            if idx >= len(row):
                continue
            price = row[idx]
            if price is None or str(price).strip() == "":
                continue
            try:
                price_f = float(price)
            except (TypeError, ValueError):
                result["skipped"] += 1
                continue
            changed = self._upsert_price(
                dataset_id=dataset_id,
                symbol=symbol,
                trade_date=trade_date,
                price=price_f,
                source=MONTHLY_PRICE_SOURCE,
            )
            if changed == "inserted":
                result["inserted"] += 1
            elif changed == "updated":
                result["updated"] += 1
            else:
                result["skipped"] += 1
            result["affected"] += 1
        return result

    def _detect_price_header(
        self, rows: list[tuple[Any, ...]]
    ) -> tuple[int, int, dict[int, str]] | None:
        for r_idx, row in enumerate(rows[:12]):
            normalized = [str(c).strip() if c is not None else "" for c in row]
            marker_count = len(set(normalized) & PRICE_HEADER_MARKERS)
            if marker_count < 2:
                continue
            symbol_cols: dict[int, str] = {}
            for idx, col in enumerate(normalized):
                mapped = PRICE_SYMBOL_MAP.get(col)
                if mapped:
                    symbol_cols[idx] = mapped
            if not symbol_cols:
                continue
            first_symbol_col = min(symbol_cols)
            date_col = max(0, first_symbol_col - 1)
            return r_idx, date_col, symbol_cols
        return None

    def _detect_balance_header(
        self, rows: list[tuple[Any, ...]]
    ) -> tuple[int, int, int, int, list[tuple[int, str]]] | None:
        for r_idx, row in enumerate(rows[:12]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if BALANCE_HEADER_NAMES["agency"] not in cells or BALANCE_HEADER_NAMES["supply_demand"] not in cells:
                continue
            agency_col = cells.index(BALANCE_HEADER_NAMES["agency"])
            update_col = cells.index(BALANCE_HEADER_NAMES["update_date"])
            sd_col = cells.index(BALANCE_HEADER_NAMES["supply_demand"])
            period_cols: list[tuple[int, str]] = []
            for idx, col in enumerate(cells):
                if idx <= sd_col or not col:
                    continue
                if re.search(r"20\d{2}(?:Q[1-4]|年)", col, re.IGNORECASE):
                    period_cols.append((idx, col))
            if period_cols:
                return r_idx, agency_col, update_col, sd_col, period_cols
        # 部分 sheet 表头被合并单元格遮挡，使用数据行形态推断。
        for r_idx, row in enumerate(rows[:12]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            agency_candidates = [idx for idx, c in enumerate(cells) if self._normalize_agency(c) in BALANCE_AGENCIES]
            sd_candidates = [idx for idx, c in enumerate(cells) if c in BALANCE_SUPPLY_DEMAND_VALUES]
            if agency_candidates and sd_candidates:
                agency_col = agency_candidates[0]
                sd_col = sd_candidates[0]
                update_col = max(0, sd_col - 1)
                header_row = max(0, r_idx - 1)
                header_cells = [str(c).strip() if c is not None else "" for c in rows[header_row]]
                period_cols = [
                    (idx, col)
                    for idx, col in enumerate(header_cells)
                    if idx > sd_col and re.search(r"20\d{2}(?:Q[1-4]|年)", col, re.IGNORECASE)
                ]
                if period_cols:
                    return header_row, agency_col, update_col, sd_col, period_cols
        return None

    def _upsert_price(
        self,
        *,
        dataset_id: int,
        symbol: str,
        trade_date: date,
        price: float,
        source: str,
    ) -> str:
        row = (
            self.db.query(PriceSeries)
            .filter(
                PriceSeries.symbol == symbol,
                PriceSeries.trade_date == trade_date,
                PriceSeries.source == source,
            )
            .first()
        )
        if row:
            if row.price == price and row.dataset_id == dataset_id:
                return "skipped"
            row.price = price
            row.dataset_id = dataset_id
            return "updated"
        self.db.add(
            PriceSeries(
                dataset_id=dataset_id,
                symbol=symbol,
                trade_date=trade_date,
                price=price,
                source=source,
            )
        )
        self.db.flush()
        return "inserted"

    def _upsert_balance(
        self,
        *,
        dataset_id: int,
        agency: str,
        snapshot_month: str,
        update_date: str,
        supply_demand: str,
        period: str,
        value: float,
        balance_gap: float | None,
    ) -> str:
        row = (
            self.db.query(BalanceForecast)
            .filter(
                BalanceForecast.agency == agency,
                BalanceForecast.snapshot_month == snapshot_month,
                BalanceForecast.supply_demand == supply_demand,
                BalanceForecast.period == period,
            )
            .first()
        )
        if row:
            if (
                row.value == value
                and row.balance_gap == balance_gap
                and row.dataset_id == dataset_id
                and row.update_date == update_date
            ):
                return "skipped"
            row.value = value
            row.balance_gap = balance_gap
            row.dataset_id = dataset_id
            row.update_date = update_date
            return "updated"
        self.db.add(
            BalanceForecast(
                dataset_id=dataset_id,
                agency=agency,
                snapshot_month=snapshot_month,
                update_date=update_date,
                supply_demand=supply_demand,
                period=period,
                value=value,
                balance_gap=balance_gap,
            )
        )
        self.db.flush()
        return "inserted"

    @staticmethod
    def _normalize_agency(value: str) -> str:
        text = value.strip()
        aliases = {"IHS": "S&P", "WoodMac": "Wood Mackenzie"}
        return aliases.get(text, text)

    @staticmethod
    def _format_update_date(value: Any, sheet_name: str) -> str:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
            return text
        year_match = re.search(r"(20\d{2})", sheet_name)
        year = year_match.group(1) if year_match else ""
        if re.fullmatch(r"\d{1,2}\.\d{1,2}", text) and year:
            month, day = text.split(".")
            return f"{year}-{int(month):02d}-{int(day):02d}"
        return text

    def _detect_factor_label_col(self, rows: list[tuple[Any, ...]]) -> int:
        """预测分析表 A 列常为空白，影响因素名称实际在 B 列（与 prediction.py 导出布局一致）。"""
        for row in rows[:40]:
            if not row:
                continue
            for idx, cell in enumerate(row[:6]):
                if cell and re.match(r"^\s*\d+\.\d+", str(cell).strip()):
                    return idx
        return 1

    def _import_factor_xlsx(self, file_path: Path) -> dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = list(wb.sheetnames)
        ds = self._create_dataset(file_path.stem, "factor", file_path, {"sheets": sheet_names}, 0)
        total_rows = 0
        imported_months: set[str] = set()

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            report_month = (
                self._parse_factor_report_month(sheet_name)
                or self._parse_factor_report_month(file_path.name)
                or self._detect_factor_report_month_from_rows(rows)
            )
            if not report_month:
                continue
            imported_months.add(report_month)
            total_rows += self._import_factor_rows(ds.id, rows, report_month)

        wb.close()
        if total_rows == 0:
            self.db.rollback()
            raise ValueError(
                json.dumps(
                    {
                        "message": "无有效因素行（请确认文件名/Sheet 含年月，且因素行含形势判断）",
                        "sheets": sheet_names,
                    },
                    ensure_ascii=False,
                )
            )

        ds.row_count = total_rows
        self.db.commit()
        return {
            "dataset_id": ds.id,
            "category": "factor",
            "rows": total_rows,
            "report_months": sorted(imported_months),
        }

    @staticmethod
    def _parse_factor_report_month(text: str) -> str | None:
        for pat in _FACTOR_MONTH_PATTERNS:
            match = pat.search(str(text))
            if match:
                year = int(match.group(1))
                month = int(match.group(2))
                if 1 <= month <= 12:
                    return f"{year}-{month:02d}"
        return None

    @staticmethod
    def _detect_factor_report_month_from_rows(rows: list[tuple[Any, ...]]) -> str | None:
        for row in rows[:20]:
            if not row:
                continue
            for cell in row[:8]:
                if cell is None:
                    continue
                parsed = DataImportService._parse_factor_report_month(str(cell))
                if parsed:
                    return parsed
        return None

    def _import_factor_rows(
        self,
        dataset_id: int,
        rows: list[tuple[Any, ...]],
        report_month: str,
    ) -> int:
        label_col = self._detect_factor_label_col(rows)
        current_category = ""
        total_rows = 0

        for row in rows:
            if not row or label_col >= len(row):
                continue
            first = str(row[label_col]).strip() if row[label_col] else ""
            if not first:
                continue
            if re.match(r"^\d+\.\S", first) and not re.match(r"^\d+\.\d+", first):
                current_category = first
                continue
            if re.match(r"^\s*\d+\.\d+", first):
                factor_name = first.strip()
                importance = 3
                for cell in row[label_col + 1 : label_col + 6]:
                    if cell in (1, 2, 3, 4, 5):
                        importance = int(cell)
                assessment = ""
                for cell in row[label_col + 6 : label_col + 14]:
                    if cell and str(cell).strip() and str(cell).strip() not in {"促涨", "持平", "促跌"}:
                        assessment = str(cell).strip()
                        break
                if not assessment:
                    for cell in row[label_col + 1 :]:
                        text = str(cell or "").strip()
                        if len(text) >= 8 and text not in {"促涨", "持平", "促跌"}:
                            assessment = text
                            break
                impact = "持平"
                for cell in row[label_col + 6 :]:
                    if cell in {"促涨", "持平", "促跌"}:
                        impact = str(cell)
                if factor_name and assessment:
                    changed = self._upsert_factor(
                        dataset_id=dataset_id,
                        report_month=report_month,
                        category=current_category,
                        factor_name=factor_name,
                        importance=importance,
                        assessment=assessment[:2000],
                        impact_direction=impact,
                    )
                    if changed != "skipped":
                        total_rows += 1
        return total_rows

    def _upsert_factor(
        self,
        *,
        dataset_id: int,
        report_month: str,
        category: str,
        factor_name: str,
        importance: int,
        assessment: str,
        impact_direction: str,
    ) -> str:
        row = (
            self.db.query(FactorAssessment)
            .filter(
                FactorAssessment.report_month == report_month,
                FactorAssessment.factor_name == factor_name,
            )
            .first()
        )
        if row:
            if (
                row.category == category
                and row.importance == importance
                and row.assessment == assessment
                and row.impact_direction == impact_direction
                and row.dataset_id == dataset_id
            ):
                return "skipped"
            row.category = category
            row.importance = importance
            row.assessment = assessment
            row.impact_direction = impact_direction
            row.dataset_id = dataset_id
            return "updated"
        self.db.add(
            FactorAssessment(
                dataset_id=dataset_id,
                report_month=report_month,
                category=category,
                factor_name=factor_name,
                importance=importance,
                assessment=assessment,
                impact_direction=impact_direction,
            )
        )
        self.db.flush()
        return "inserted"

    def _parse_date(self, value: Any):
        if value is None:
            return None
        if hasattr(value, "date"):
            return value.date()
        if isinstance(value, int) and 19000101 <= value <= 21001231:
            return datetime.strptime(str(value), "%Y%m%d").date()
        s = str(value).strip()
        if re.fullmatch(r"\d{8}", s):
            return datetime.strptime(s, "%Y%m%d").date()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
        return None

    def quality_check(self, dataset_id: int) -> dict[str, Any]:
        ds = self.db.get(Dataset, dataset_id)
        if not ds:
            raise ValueError("Dataset not found")
        issues = []
        if ds.category == "price":
            count = self.db.query(PriceSeries).filter(PriceSeries.dataset_id == dataset_id).count()
            if count == 0:
                issues.append("No price rows imported")
            symbols = (
                self.db.query(PriceSeries.symbol)
                .filter(PriceSeries.dataset_id == dataset_id)
                .distinct()
                .all()
            )
            if len(symbols) < 2:
                issues.append("Less than 2 price symbols detected")
        return {"dataset_id": dataset_id, "issues": issues, "passed": len(issues) == 0}

    def clear_data(self, category: str | None = None) -> dict[str, int]:
        """清空已导入的业务数据。category 为空时清空全部。"""
        counts: dict[str, int] = {}

        if category in (None, "price"):
            q = self.db.query(PriceSeries)
            if category == "price":
                price_ds = [r.id for r in self.db.query(Dataset.id).filter(Dataset.category == "price").all()]
                q = q.filter(PriceSeries.dataset_id.in_(price_ds)) if price_ds else q.filter(False)
            counts["price_series"] = q.delete(synchronize_session=False)

        if category in (None, "balance"):
            q = self.db.query(BalanceForecast)
            if category == "balance":
                balance_ds = [r.id for r in self.db.query(Dataset.id).filter(Dataset.category == "balance").all()]
                q = q.filter(BalanceForecast.dataset_id.in_(balance_ds)) if balance_ds else q.filter(False)
            counts["balance_forecast"] = q.delete(synchronize_session=False)

        if category in (None, "factor"):
            q = self.db.query(FactorAssessment)
            if category == "factor":
                factor_ds = [r.id for r in self.db.query(Dataset.id).filter(Dataset.category == "factor").all()]
                q = q.filter(FactorAssessment.dataset_id.in_(factor_ds)) if factor_ds else q.filter(False)
            counts["factor_assessment"] = q.delete(synchronize_session=False)

        ds_q = self.db.query(Dataset)
        if category:
            ds_q = ds_q.filter(Dataset.category == category)
        counts["datasets"] = ds_q.delete(synchronize_session=False)

        self.db.commit()
        return counts

    def seed_sample_data(self, sample_dir: Path) -> list[dict[str, Any]]:
        results = []
        if not sample_dir.exists():
            return results
        for path in sorted(sample_dir.rglob("*.xlsx")):
            try:
                results.append(self.import_file(path))
            except Exception as exc:
                results.append({"file": str(path), "error": str(exc)})
        return results
