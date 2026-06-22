#!/usr/bin/env python3
"""对账样例 xlsx 与数据库覆盖情况。"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import func

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import SessionLocal, init_db  # noqa: E402
from app.models.balance_forecast import BalanceForecast  # noqa: E402
from app.models.price_series import PriceSeries  # noqa: E402

PRICE_SHEET_RE = re.compile(r"^(\d{4})年(\d{1,2})月$")
BALANCE_SHEET_RE = re.compile(r"^(\d{4})(\d{2})$")
YUEBAO = ROOT / "yuebao"


def audit_price(db) -> dict:
    path = YUEBAO / "原油价格表. 20260603.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    expected = []
    for name in wb.sheetnames:
        m = PRICE_SHEET_RE.match(name.strip())
        if m:
            expected.append(f"{m.group(1)}-{int(m.group(2)):02d}")
    wb.close()

    actual = {
        f"{y}-{int(m):02d}"
        for y, m in db.query(
            func.strftime("%Y", PriceSeries.trade_date),
            func.strftime("%m", PriceSeries.trade_date),
        )
        .filter(PriceSeries.source == "CNEEI")
        .distinct()
        .all()
        if y and m
    }
    missing = sorted(set(expected) - actual)
    return {"expected": len(expected), "actual": len(actual), "missing_months": missing}


def audit_balance(db) -> dict:
    path = YUEBAO / "供需平衡表.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    expected = [name for name in wb.sheetnames if BALANCE_SHEET_RE.fullmatch(name.strip())]
    wb.close()

    actual = {
        r[0]
        for r in db.query(BalanceForecast.snapshot_month)
        .filter(BalanceForecast.snapshot_month != "")
        .distinct()
        .all()
    }
    missing = sorted(
        {f"{n[:4]}-{n[4:]}" for n in expected} - actual,
        key=lambda x: x,
    )
    dirty = (
        db.query(BalanceForecast)
        .filter(BalanceForecast.period.like("col_%"))
        .count()
    )
    return {
        "expected_sheets": len(expected),
        "actual_snapshots": len(actual),
        "missing_snapshots": missing,
        "dirty_period_rows": dirty,
    }


def main() -> None:
    init_db()
    db = SessionLocal()
    try:
        price = audit_price(db)
        balance = audit_balance(db)
        print("=== 价格表 ===")
        print(price)
        print("=== 供需平衡表 ===")
        print(balance)
        ok = not price["missing_months"] and not balance["missing_snapshots"] and balance["dirty_period_rows"] == 0
        print("PASS" if ok else "FAIL")
        sys.exit(0 if ok else 1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
